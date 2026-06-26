import sqlite3
import pandas as pd
import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure Vietnamese Unicode characters are printed properly on Windows console
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# Path configurations
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DB_PATH = os.path.join(PROJECT_ROOT, "hotel_warehouse.db")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "csv_exports")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "hotel_data_warehouse.xlsx")

def build_flat_hotels_df(conn):
    """
    Builds a flat hotel dataframe (1 row per hotel) with aggregated prices and room lists.
    """
    cursor = conn.cursor()
    
    # 1. Query base hotel data with location and review info
    hotels_query = """
        SELECT 
            h.hotel_id, 
            h.hotel_name, 
            h.property_type,
            h.stars_rating, 
            h.review_count, 
            h.popular_facilities, 
            h.description, 
            h.hotel_policies,
            l.address, 
            l.whats_nearby,
            l.top_attractions, 
            l.natural_beauty, 
            l.nearest_airports, 
            l.restaurants_cafes, 
            l.beaches_in_area, 
            l.public_transport, 
            r.average_score,
            r.staff_score, 
            r.facilities_score, 
            r.cleanliness_score, 
            r.comfort_score, 
            r.value_for_money_score, 
            r.location_score, 
            r.free_wifi_score
        FROM Dim_Hotel h
        LEFT JOIN Dim_Location l ON h.hotel_id = l.hotel_id
        LEFT JOIN Dim_Review r ON h.hotel_id = r.hotel_id
    """
    
    cursor.execute(hotels_query)
    hotels = cursor.fetchall()
    hotel_cols = [desc[0] for desc in cursor.description]
    
    flat_data = []
    
    for hotel in hotels:
        hotel_dict = dict(zip(hotel_cols, hotel))
        hotel_id = hotel_dict["hotel_id"]
        
        # 2. Query room names and prices
        rooms_query = """
            SELECT 
                rt.room_type_name, 
                fb.current_price
            FROM Fact_Booking fb
            JOIN Dim_RoomType rt ON fb.room_type_id = rt.room_type_id
            WHERE fb.hotel_id = ?
        """
        cursor.execute(rooms_query, (hotel_id,))
        room_prices = cursor.fetchall()
        
        # 3. Query distinct booking checkin/checkout dates
        dates_query = """
            SELECT DISTINCT checkin_date, checkout_date
            FROM Fact_Booking
            WHERE hotel_id = ? AND checkin_date IS NOT NULL
        """
        cursor.execute(dates_query, (hotel_id,))
        booking_dates = cursor.fetchall()
        
        if booking_dates:
            checkin_str = ", ".join(sorted(list(set([d[0] for d in booking_dates if d[0]]))))
            checkout_str = ", ".join(sorted(list(set([d[1] for d in booking_dates if d[1]]))))
        else:
            checkin_str = ""
            checkout_str = ""
            
        if room_prices:
            prices = [rp[1] for rp in room_prices if rp[1] is not None and rp[1] > 0]
            min_price = min(prices) if prices else 0.0
            max_price = max(prices) if prices else 0.0
            room_list_str = "; ".join([f"{rp[0]} ({int(rp[1]):,} VND)" if rp[1] is not None else rp[0] for rp in room_prices])
        else:
            min_price = 0.0
            max_price = 0.0
            room_list_str = ""
            
        hotel_dict["ngay_checkin"] = checkin_str
        hotel_dict["ngay_checkout"] = checkout_str
        hotel_dict["gia_thap_nhat"] = min_price
        hotel_dict["gia_cao_nhat"] = max_price
        hotel_dict["danh_sach_phong_va_gia"] = room_list_str
        
        flat_data.append(hotel_dict)
        
    df = pd.DataFrame(flat_data)
    
    # Reorder or rename flat columns to look clean in Vietnamese
    rename_cols = {
        "hotel_id": "Mã khách sạn",
        "hotel_name": "Tên khách sạn",
        "property_type": "Loại hình",
        "stars_rating": "Số sao",
        "review_count": "Số lượng đánh giá",
        "popular_facilities": "Tiện ích nổi bật",
        "description": "Mô tả",
        "hotel_policies": "Chính sách khách sạn",
        "address": "Địa chỉ",
        "whats_nearby": "Địa điểm lân cận",
        "top_attractions": "Điểm du lịch hàng đầu",
        "natural_beauty": "Cảnh quan thiên nhiên",
        "nearest_airports": "Sân bay gần nhất",
        "restaurants_cafes": "Nhà hàng & Quán cafe",
        "beaches_in_area": "Bãi biển gần đây",
        "public_transport": "Phương tiện công cộng",
        "average_score": "Điểm trung bình",
        "staff_score": "Điểm phục vụ",
        "facilities_score": "Điểm tiện nghi",
        "cleanliness_score": "Điểm sạch sẽ",
        "comfort_score": "Điểm thoải mái",
        "value_for_money_score": "Điểm đáng giá tiền",
        "location_score": "Điểm vị trí",
        "free_wifi_score": "Điểm Wifi miễn phí",
        "ngay_checkin": "Ngày Check-in",
        "ngay_checkout": "Ngày Check-out",
        "gia_thap_nhat": "Giá thấp nhất (VND)",
        "gia_cao_nhat": "Giá cao nhất (VND)",
        "danh_sach_phong_va_gia": "Danh sách phòng & Giá"
    }
    
    # We select columns in a neat logical order
    col_order = [
        "hotel_id", "hotel_name", "property_type", "stars_rating", "average_score", "review_count",
        "gia_thap_nhat", "gia_cao_nhat", "address", "ngay_checkin", "ngay_checkout",
        "popular_facilities", "description", "hotel_policies", "whats_nearby", "top_attractions",
        "natural_beauty", "nearest_airports", "restaurants_cafes", "beaches_in_area", "public_transport",
        "staff_score", "facilities_score", "cleanliness_score", "comfort_score", "value_for_money_score",
        "location_score", "free_wifi_score", "danh_sach_phong_va_gia"
    ]
    
    df = df[col_order].rename(columns=rename_cols)
    return df

def export_warehouse_to_xlsx():
    if not os.path.exists(DB_PATH):
        print(f"Không tìm thấy file database tại {DB_PATH}")
        return

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    print("="*60)
    print("XUẤT DỮ LIỆU WAREHOUSE RA EXCEL (.xlsx) CHUYÊN NGHIỆP")
    print("="*60)

    conn = sqlite3.connect(DB_PATH)
    
    # Load all tables
    print("Đang đọc dữ liệu các bảng từ SQLite...")
    dim_hotel = pd.read_sql_query("SELECT * FROM Dim_Hotel", conn)
    dim_location = pd.read_sql_query("SELECT * FROM Dim_Location", conn)
    dim_review = pd.read_sql_query("SELECT * FROM Dim_Review", conn)
    dim_roomtype = pd.read_sql_query("SELECT * FROM Dim_RoomType", conn)
    fact_booking = pd.read_sql_query("SELECT * FROM Fact_Booking", conn)
    
    # Build flat consolidated view
    print("Đang tổng hợp thông tin chi tiết từng khách sạn...")
    flat_hotels = build_flat_hotels_df(conn)
    
    conn.close()

    # Calculate summary metrics for the Dashboard
    total_hotels = len(dim_hotel)
    total_rooms = len(dim_roomtype)
    total_bookings = len(fact_booking)
    avg_score = dim_review['average_score'].mean() if len(dim_review) > 0 else 0.0

    print(f"Tổng hợp thành công:")
    print(f"  - Số lượng khách sạn: {total_hotels}")
    print(f"  - Số lượng loại phòng: {total_rooms}")
    print(f"  - Số lượng bản ghi giá/booking: {total_bookings}")
    print(f"  - Điểm đánh giá trung bình: {avg_score:.2f}")

    # Write to Excel
    print(f"\nĐang ghi dữ liệu vào file Excel: {OUTPUT_FILE}")
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Create writer empty sheet
        workbook = writer.book
        
        # 1. Create Dashboard/Readme sheet first
        ws_dashboard = workbook.create_sheet(title="Giới thiệu & Thống kê")
        
        # Write pandas tables to sheets
        flat_hotels.to_excel(writer, sheet_name="Tổng hợp khách sạn", index=False)
        dim_hotel.to_excel(writer, sheet_name="Dim_Hotel", index=False)
        dim_location.to_excel(writer, sheet_name="Dim_Location", index=False)
        dim_review.to_excel(writer, sheet_name="Dim_Review", index=False)
        dim_roomtype.to_excel(writer, sheet_name="Dim_RoomType", index=False)
        fact_booking.to_excel(writer, sheet_name="Fact_Booking", index=False)
        
        # Populate Dashboard / Readme
        ws_dashboard.views.sheetView[0].showGridLines = True
        
        # Styles
        font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
        font_subtitle = Font(name="Segoe UI", size=11, italic=True, color="595959")
        font_section = Font(name="Segoe UI", size=12, bold=True, color="1F4E78")
        font_bold = Font(name="Segoe UI", size=10, bold=True)
        font_regular = Font(name="Segoe UI", size=10)
        
        fill_navy_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_light_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Add Title
        ws_dashboard["B2"] = "HỆ THỐNG KHO DỮ LIỆU KHÁCH SẠN ĐÀ NẴNG (BOOKING.COM)"
        ws_dashboard["B2"].font = font_title
        ws_dashboard["B3"] = f"Dữ liệu được cào tự động qua Firecrawl API - Cập nhật: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_dashboard["B3"].font = font_subtitle
        
        # Add Summary Statistics Card
        ws_dashboard["B5"] = "THỐNG KÊ TỔNG QUAN"
        ws_dashboard["B5"].font = font_section
        
        stats_headers = ["Chỉ số", "Giá trị"]
        stats_data = [
            ["Tổng số khách sạn đã cào", total_hotels],
            ["Tổng số loại phòng phân tích", total_rooms],
            ["Tổng số điểm dữ liệu giá (Booking)", total_bookings],
            ["Điểm đánh giá trung bình", round(avg_score, 2)]
        ]
        
        # Write headers
        for col_idx, header in enumerate(stats_headers, start=2):
            cell = ws_dashboard.cell(row=6, column=col_idx, value=header)
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            cell.fill = fill_navy_header
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        # Write data
        for row_idx, row_data in enumerate(stats_data, start=7):
            for col_idx, value in enumerate(row_data, start=2):
                cell = ws_dashboard.cell(row=row_idx, column=col_idx, value=value)
                cell.font = font_regular
                cell.border = thin_border
                if col_idx == 2:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.font = font_bold
                    cell.alignment = Alignment(horizontal="right")
                    if isinstance(value, float):
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '#,##0'
                        
        # Add Table Information
        ws_dashboard["B13"] = "DANH SÁCH BẢNG DỮ LIỆU (SHEETS)"
        ws_dashboard["B13"].font = font_section
        
        sheet_info_headers = ["Tên Sheet", "Loại thông tin", "Mô tả chi tiết"]
        sheet_info_data = [
            ["Tổng hợp khách sạn", "Consolidated Flat View", "Thông tin gộp 1 khách sạn/dòng (tên, sao, điểm trung bình, địa chỉ, khoảng giá, danh sách phòng)"],
            ["Dim_Hotel", "Dimension Table", "Thông tin cơ bản về khách sạn: Mã khách sạn, Tên, Số sao, Lượt đánh giá, Tiện ích, Chính sách chung"],
            ["Dim_Location", "Dimension Table", "Thông tin địa lý: Địa chỉ chi tiết, Các điểm tham quan gần đó, Sân bay, Bãi biển, Xe buýt..."],
            ["Dim_Review", "Dimension Table", "Bảng điểm đánh giá chi tiết: Nhân viên, Sạch sẽ, Tiện nghi, Thoải mái, Giá tiền, Vị trí, Wifi..."],
            ["Dim_RoomType", "Dimension Table", "Các loại phòng có tại khách sạn: Tên loại phòng, Loại giường, Số khách tối đa"],
            ["Fact_Booking", "Fact Table", "Các sự kiện đặt phòng: Giá gốc, Giá hiện tại sau giảm giá, Chiết khấu, Thuế phí, Ngày checkin/checkout"]
        ]
        
        # Write sheet info headers
        for col_idx, header in enumerate(sheet_info_headers, start=2):
            cell = ws_dashboard.cell(row=14, column=col_idx, value=header)
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            cell.fill = fill_navy_header
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        # Write sheet info data
        for row_idx, row_data in enumerate(sheet_info_data, start=15):
            for col_idx, value in enumerate(row_data, start=2):
                cell = ws_dashboard.cell(row=row_idx, column=col_idx, value=value)
                cell.font = font_regular
                cell.border = thin_border
                if col_idx == 2:
                    cell.font = font_bold
                    cell.fill = fill_light_blue
                cell.alignment = Alignment(horizontal="left")
                
        # Adjust dashboard column widths
        ws_dashboard.column_dimensions['A'].width = 3
        ws_dashboard.column_dimensions['B'].width = 30
        ws_dashboard.column_dimensions['C'].width = 25
        ws_dashboard.column_dimensions['D'].width = 90
        
        # Style other data sheets
        for sheet_name in workbook.sheetnames:
            if sheet_name == "Giới thiệu & Thống kê":
                continue
                
            ws = workbook[sheet_name]
            ws.views.sheetView[0].showGridLines = True
            
            # Format Headers
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                cell.fill = fill_navy_header
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
            ws.row_dimensions[1].height = 28
            
            # Formatting cells and aligning columns
            for row_idx in range(2, ws.max_row + 1):
                # Zebra striping
                fill = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
                
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = font_regular
                    cell.border = thin_border
                    if fill.fill_type:
                        cell.fill = fill
                        
                    # Get column header
                    col_header = ws.cell(row=1, column=col_idx).value
                    val = cell.value
                    
                    # Formatting values
                    if val is not None:
                        # Price columns formatting
                        if any(term in str(col_header).lower() for term in ["price", "discount", "giá"]):
                            try:
                                cell.value = float(val)
                                cell.number_format = '#,##0'
                                cell.alignment = Alignment(horizontal="right")
                            except:
                                cell.alignment = Alignment(horizontal="left")
                        # Score columns formatting
                        elif any(term in str(col_header).lower() for term in ["score", "điểm"]):
                            try:
                                cell.value = float(val)
                                cell.number_format = '0.0'
                                cell.alignment = Alignment(horizontal="right")
                            except:
                                cell.alignment = Alignment(horizontal="left")
                        # Counts, id and regular numbers
                        elif any(term in str(col_header).lower() for term in ["count", "max_guests", "booking_id", "review_id", "taxes", "thuế", "số lượng"]):
                            try:
                                cell.value = int(val)
                                cell.number_format = '#,##0'
                                cell.alignment = Alignment(horizontal="right")
                            except:
                                cell.alignment = Alignment(horizontal="center")
                        # Dates alignment
                        elif any(term in str(col_header).lower() for term in ["date", "extracted_at", "ngày"]):
                            cell.alignment = Alignment(horizontal="center")
                        # ID columns alignment
                        elif "id" in str(col_header).lower() or "mã" in str(col_header).lower():
                            cell.alignment = Alignment(horizontal="center")
                        # Default alignment for long text vs short text
                        else:
                            if len(str(val)) > 30:
                                cell.alignment = Alignment(horizontal="left", wrap_text=True)
                            else:
                                cell.alignment = Alignment(horizontal="left")
                    else:
                        cell.alignment = Alignment(horizontal="left")

            # Auto-fit columns with safety margin
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                
                # Check column header length
                header_val = ws.cell(row=1, column=col[0].column).value
                header_len = len(str(header_val)) if header_val else 10
                
                # Scan data values (sample first 100 rows for performance if very large)
                sample_rows = col[:150]
                for cell in sample_rows:
                    if cell.value:
                        # Avoid overly long width for description columns
                        val_len = len(str(cell.value))
                        if val_len > max_len:
                            max_len = val_len
                            
                # Determine final width
                final_width = max(max_len + 3, header_len + 5)
                # Cap the column width for readable layout (especially description/facilities)
                if final_width > 50:
                    final_width = 50
                ws.column_dimensions[col_letter].width = final_width
                
    print(f"\n-> Xuất file Excel thành công tại: {OUTPUT_FILE}")
    print("="*60)

if __name__ == "__main__":
    export_warehouse_to_xlsx()
